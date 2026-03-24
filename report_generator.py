"""
report_generator.py
Professional PDF report + Excel session log for the Keyence IV3 Camera Calculator.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd

# ── ReportLab imports ────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus.flowables import Flowable

# ── Brand colours ────────────────────────────────────────────────────────────
C_NAVY   = colors.HexColor("#0d1b2a")
C_BLUE   = colors.HexColor("#1a73e8")
C_BLUE_L = colors.HexColor("#e8f0fd")
C_GREEN  = colors.HexColor("#0d9c50")
C_GREEN_L= colors.HexColor("#e6f4ea")
C_ORANGE = colors.HexColor("#e87e1a")
C_RED    = colors.HexColor("#d93025")
C_RED_L  = colors.HexColor("#fce8e6")
C_GREY   = colors.HexColor("#f5f5f5")
C_DKGREY = colors.HexColor("#555555")
C_LINE   = colors.HexColor("#d0d0d0")
C_WHITE  = colors.white

PAGE_W, PAGE_H = A4
MARGIN = 18 * mm


# ════════════════════════════════════════════════════════════════════════════
# CUSTOM FLOWABLES
# ════════════════════════════════════════════════════════════════════════════

class HeaderBand(Flowable):
    """Full-width dark navy header band with title and subtitle."""
    def __init__(self, title: str, subtitle: str, width: float):
        super().__init__()
        self.title    = title
        self.subtitle = subtitle
        self.band_w   = width
        self.band_h   = 28 * mm

    def wrap(self, *_):
        return self.band_w, self.band_h

    def draw(self):
        c = self.canv
        # Background
        c.setFillColor(C_NAVY)
        c.rect(0, 0, self.band_w, self.band_h, fill=1, stroke=0)
        # Accent stripe
        c.setFillColor(C_BLUE)
        c.rect(0, 0, 5, self.band_h, fill=1, stroke=0)
        # Title
        c.setFillColor(C_WHITE)
        c.setFont("Helvetica-Bold", 18)
        c.drawString(12, self.band_h - 18, self.title)
        # Subtitle
        c.setFillColor(colors.HexColor("#aac8f0"))
        c.setFont("Helvetica", 9)
        c.drawString(12, 8, self.subtitle)
        # Right logo placeholder
        c.setFillColor(colors.HexColor("#1e3a5f"))
        c.roundRect(self.band_w - 52, 6, 46, 16, 3, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#aac8f0"))
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(self.band_w - 29, 11, "KEYENCE IV3")


class SectionTitle(Flowable):
    """Coloured section divider with label."""
    def __init__(self, text: str, width: float, color=C_BLUE):
        super().__init__()
        self.text  = text
        self.w     = width
        self.color = color
        self.h     = 8 * mm

    def wrap(self, *_):
        return self.w, self.h

    def draw(self):
        c = self.canv
        c.setFillColor(self.color)
        c.rect(0, self.h - 2, self.w, 2, fill=1, stroke=0)
        c.rect(0, self.h - 2, 30, 2, fill=1, stroke=0)
        c.setFillColor(self.color)
        c.roundRect(0, 0, len(self.text) * 5.8 + 16, self.h - 4, 2, fill=1, stroke=0)
        c.setFillColor(C_WHITE)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(8, 3, self.text.upper())


class InstallationDrawing(Flowable):
    """
    Engineering-style mounting sketch:
    camera body → dashed FOV lines → workpiece line
    with dimension arrows.
    """
    def __init__(
        self,
        width: float,
        height: float,
        fov_h: float,
        fov_v: float,
        distance: float,
        roi_w: Optional[float] = None,
        roi_h: Optional[float] = None,
    ):
        super().__init__()
        self.dw       = width
        self.dh       = height
        self.fov_h    = fov_h
        self.fov_v    = fov_v
        self.distance = distance
        self.roi_w    = roi_w
        self.roi_h    = roi_h

    def wrap(self, *_):
        return self.dw, self.dh

    def draw(self):
        c   = self.canv
        dw  = self.dw
        dh  = self.dh

        # Layout constants
        cam_x      = dw / 2
        cam_top_y  = dh - 12 * mm
        cam_bot_y  = cam_top_y - 10 * mm
        part_y     = 8 * mm
        half_fov   = (self.fov_h / self.fov_v) * (cam_top_y - part_y) * 0.35
        half_fov   = min(half_fov, dw * 0.38)

        # ── Background ───────────────────────────────────────────────────────
        c.setFillColor(colors.HexColor("#fafbfc"))
        c.setStrokeColor(C_LINE)
        c.roundRect(0, 0, dw, dh, 3, fill=1, stroke=1)

        # Title strip
        c.setFillColor(colors.HexColor("#f0f4f8"))
        c.rect(0, dh - 6 * mm, dw, 6 * mm, fill=1, stroke=0)
        c.setFillColor(C_NAVY)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(dw / 2, dh - 4 * mm, "INSTALLATION DRAWING — SIDE VIEW (not to scale)")

        # ── FOV cone (dashed) ────────────────────────────────────────────────
        c.setStrokeColor(C_BLUE)
        c.setLineWidth(0.6)
        c.setDash(4, 3)
        c.line(cam_x, cam_bot_y, cam_x - half_fov, part_y)
        c.line(cam_x, cam_bot_y, cam_x + half_fov, part_y)
        c.setDash()  # reset

        # FOV fill
        from reportlab.graphics.shapes import Polygon
        c.setFillColor(colors.HexColor("#e8f0fd"))
        p = c.beginPath()
        p.moveTo(cam_x, cam_bot_y)
        p.lineTo(cam_x - half_fov, part_y)
        p.lineTo(cam_x + half_fov, part_y)
        p.close()
        c.drawPath(p, fill=1, stroke=0)

        # ── ROI overlay ──────────────────────────────────────────────────────
        if self.roi_w and self.roi_h and self.fov_h > 0:
            roi_scale  = self.roi_w / self.fov_h
            roi_half   = half_fov * roi_scale
            roi_y_frac = 0.3
            roi_y_top  = part_y + (cam_bot_y - part_y) * roi_y_frac + 6 * mm
            roi_y_bot  = part_y + 2 * mm
            c.setFillColor(colors.HexColor("#fce8e6"))
            c.setStrokeColor(C_RED)
            c.setLineWidth(0.8)
            c.rect(cam_x - roi_half, roi_y_bot, roi_half * 2, roi_y_top - roi_y_bot, fill=1, stroke=1)
            c.setFillColor(C_RED)
            c.setFont("Helvetica-Bold", 6)
            c.drawCentredString(cam_x, (roi_y_bot + roi_y_top) / 2, f"ROI  {self.roi_w:.0f}×{self.roi_h:.0f} mm")

        # ── Camera body ──────────────────────────────────────────────────────
        cb_w, cb_h = 14 * mm, 10 * mm
        c.setFillColor(C_NAVY)
        c.setStrokeColor(C_NAVY)
        c.setLineWidth(0.5)
        c.roundRect(cam_x - cb_w / 2, cam_bot_y, cb_w, cb_h, 2, fill=1, stroke=1)
        # Lens circle
        c.setFillColor(colors.HexColor("#3a6a9a"))
        c.circle(cam_x, cam_bot_y + cb_h / 2, 3 * mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#7ab0e0"))
        c.circle(cam_x, cam_bot_y + cb_h / 2, 1.5 * mm, fill=1, stroke=0)
        # Label
        c.setFillColor(C_WHITE)
        c.setFont("Helvetica-Bold", 5.5)
        c.drawCentredString(cam_x + 3 * mm, cam_bot_y + 2, "IV3")

        # ── Workpiece / part line ────────────────────────────────────────────
        c.setStrokeColor(C_DKGREY)
        c.setFillColor(C_DKGREY)
        c.setLineWidth(2)
        c.line(8 * mm, part_y, dw - 8 * mm, part_y)
        c.setFont("Helvetica", 6)
        c.drawString(dw - 24 * mm, part_y + 1.5 * mm, "PART / WORKPIECE")

        # ── Dimension: mounting distance ─────────────────────────────────────
        dim_x   = dw - 10 * mm
        c.setStrokeColor(C_ORANGE)
        c.setFillColor(C_ORANGE)
        c.setLineWidth(0.7)
        c.setDash(2, 2)
        c.line(cam_x + cb_w / 2 + 2 * mm, cam_bot_y, dim_x, cam_bot_y)
        c.line(cam_x + cb_w / 2 + 2 * mm, part_y, dim_x, part_y)
        c.setDash()
        # Arrow line
        c.setLineWidth(1)
        c.line(dim_x, part_y, dim_x, cam_bot_y)
        # Arrowheads
        arr = 1.5 * mm
        c.line(dim_x, cam_bot_y, dim_x - arr, cam_bot_y + arr)
        c.line(dim_x, cam_bot_y, dim_x + arr, cam_bot_y + arr)
        c.line(dim_x, part_y,   dim_x - arr, part_y - arr)
        c.line(dim_x, part_y,   dim_x + arr, part_y - arr)
        # Dim label
        mid_y = (cam_bot_y + part_y) / 2
        c.setFillColor(C_WHITE)
        lbl_w = 18 * mm
        c.roundRect(dim_x - lbl_w / 2 + 0.5 * mm, mid_y - 3.5 * mm, lbl_w, 7 * mm, 2, fill=1, stroke=0)
        c.setFillColor(C_ORANGE)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(dim_x + 0.5 * mm, mid_y + 1, f"{self.distance:.0f} mm")
        c.setFont("Helvetica", 5.5)
        c.drawCentredString(dim_x + 0.5 * mm, mid_y - 3, "MOUNT. DIST.")

        # ── Dimension: FOV width ─────────────────────────────────────────────
        fov_dim_y = part_y - 8 * mm
        c.setStrokeColor(C_BLUE)
        c.setFillColor(C_BLUE)
        c.setLineWidth(0.7)
        c.setDash(2, 2)
        c.line(cam_x - half_fov, part_y, cam_x - half_fov, fov_dim_y)
        c.line(cam_x + half_fov, part_y, cam_x + half_fov, fov_dim_y)
        c.setDash()
        c.setLineWidth(1)
        c.line(cam_x - half_fov, fov_dim_y, cam_x + half_fov, fov_dim_y)
        c.line(cam_x - half_fov, fov_dim_y, cam_x - half_fov + arr, fov_dim_y + arr)
        c.line(cam_x - half_fov, fov_dim_y, cam_x - half_fov + arr, fov_dim_y - arr)
        c.line(cam_x + half_fov, fov_dim_y, cam_x + half_fov - arr, fov_dim_y + arr)
        c.line(cam_x + half_fov, fov_dim_y, cam_x + half_fov - arr, fov_dim_y - arr)
        c.setFillColor(C_WHITE)
        c.roundRect(cam_x - 11 * mm, fov_dim_y - 3.5 * mm, 22 * mm, 7 * mm, 2, fill=1, stroke=0)
        c.setFillColor(C_BLUE)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(cam_x, fov_dim_y + 1, f"{self.fov_h:.1f} mm (H)")
        c.setFont("Helvetica", 5.5)
        c.drawCentredString(cam_x, fov_dim_y - 3, "FOV WIDTH")


# ════════════════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def _styles():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "Title2", parent=base["Normal"],
            fontName="Helvetica-Bold", fontSize=13, textColor=C_NAVY,
            spaceAfter=2,
        ),
        "subtitle": ParagraphStyle(
            "Sub", parent=base["Normal"],
            fontName="Helvetica", fontSize=8, textColor=C_DKGREY,
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "Body2", parent=base["Normal"],
            fontName="Helvetica", fontSize=8, textColor=colors.black,
            leading=12,
        ),
        "small": ParagraphStyle(
            "Small", parent=base["Normal"],
            fontName="Helvetica", fontSize=7, textColor=C_DKGREY,
        ),
        "bold": ParagraphStyle(
            "Bold2", parent=base["Normal"],
            fontName="Helvetica-Bold", fontSize=8, textColor=colors.black,
        ),
        "verdict_ok": ParagraphStyle(
            "VerdOK", parent=base["Normal"],
            fontName="Helvetica-Bold", fontSize=10, textColor=C_GREEN,
            backColor=C_GREEN_L, borderPad=6,
        ),
        "verdict_bad": ParagraphStyle(
            "VerdBAD", parent=base["Normal"],
            fontName="Helvetica-Bold", fontSize=10, textColor=C_RED,
            backColor=C_RED_L, borderPad=6,
        ),
    }


def _metric_table(rows: list[tuple[str, str]], col_w=None) -> Table:
    """Two-column label/value table styled as metric cards."""
    usable = PAGE_W - 2 * MARGIN
    if col_w is None:
        col_w = [usable * 0.52, usable * 0.48]
    data = [["PARAMETER", "VALUE"]] + list(rows)
    t = Table(data, colWidths=col_w)
    style = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), C_WHITE),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 7),
        ("TOPPADDING", (0, 0), (-1, 0), 4),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 4),
        # Body
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",   (1, 1), (1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.4, C_LINE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_GREY]),
        ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
    ]
    t.setStyle(TableStyle(style))
    return t


def _four_metric_row(items: list[tuple[str, str, str]]) -> Table:
    """Four metric cards in a horizontal row. items = [(label, value, unit), ...]"""
    usable = PAGE_W - 2 * MARGIN
    cw     = usable / len(items)
    data   = [[Paragraph(
        f'<font size="6" color="#666666">{lbl.upper()}</font><br/>'
        f'<font size="11"><b>{val}</b></font>'
        f'<font size="7" color="#888888"> {unit}</font>',
        getSampleStyleSheet()["Normal"]
    ) for lbl, val, unit in items]]
    t = Table(data, colWidths=[cw] * len(items))
    style = [
        ("BOX",         (0, 0), (-1, -1), 0.4, C_LINE),
        ("LINEAFTER",   (0, 0), (-2, -1), 0.4, C_LINE),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND",  (0, 0), (-1, -1), C_GREY),
        ("LINEBEFORE",  (0, 0), (0, -1), 3, C_BLUE),
    ]
    t.setStyle(TableStyle(style))
    return t


# ════════════════════════════════════════════════════════════════════════════
# MAIN PDF BUILDER
# ════════════════════════════════════════════════════════════════════════════

def generate_pdf_report(params: dict) -> bytes:
    """
    Build a professional PDF report.

    params keys:
        camera_name, camera_specs  (dict)
        distance, fov_x, fov_y
        mpp_h, ppm_h, mpp_v, ppm_v
        roi_enabled, roi_w, roi_h, roi_px_h, roi_px_v, roi_coverage
        detection_enabled, feature_size, detect_px, detect_verdict  ('ok'|'marginal'|'fail')
        detect_min_1px, detect_min_3px
        operator, project, notes
    """
    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN,  bottomMargin=MARGIN,
        title="Keyence Camera Installation Report",
        author=params.get("operator", "Dejan Rožič"),
    )
    S     = _styles()
    story = []
    usable_w = PAGE_W - 2 * MARGIN
    now   = datetime.now().strftime("%d %b %Y  %H:%M")

    # ── COVER BAND ────────────────────────────────────────────────────────────
    story.append(HeaderBand(
        "Camera Installation Report",
        f"Keyence IV3 Series  ·  Generated: {now}  ·  {params.get('project', '')}",
        usable_w,
    ))
    story.append(Spacer(1, 6 * mm))

    # Summary row (camera / operator / date / project)
    meta = [
        ("Camera Model",  params["camera_name"],  ""),
        ("Operator",      params.get("operator", "—"), ""),
        ("Project",       params.get("project", "—"),  ""),
        ("Date",          now.split()[0] + " " + now.split()[1] + " " + now.split()[2], ""),
    ]
    story.append(_four_metric_row(meta))
    story.append(Spacer(1, 6 * mm))

    # ── SECTION 1: FOV & RESOLUTION ───────────────────────────────────────────
    story.append(SectionTitle("1. Field of View & Resolution", usable_w))
    story.append(Spacer(1, 3 * mm))

    fov_metrics = [
        ("Mounting Distance", f"{params['distance']:.0f}", "mm"),
        ("FOV Horizontal",    f"{params['fov_x']:.1f}",   "mm"),
        ("FOV Vertical",      f"{params['fov_y']:.1f}",   "mm"),
        ("Aspect Ratio",      f"{params['fov_x']/params['fov_y']:.2f}", ": 1"),
    ]
    story.append(_four_metric_row(fov_metrics))
    story.append(Spacer(1, 3 * mm))

    res_rows = [
        ("FOV Horizontal",         f"{params['fov_x']:.2f} mm"),
        ("FOV Vertical",           f"{params['fov_y']:.2f} mm"),
        ("Sensor Resolution",      f"{params.get('res_h', 1280)} × {params.get('res_v', 960)} px"),
        ("Resolution (H)",         f"{params['mpp_h']:.4f} mm/px  |  {params['ppm_h']:.2f} px/mm"),
        ("Resolution (V)",         f"{params['mpp_v']:.4f} mm/px  |  {params['ppm_v']:.2f} px/mm"),
        ("Mounting Distance",      f"{params['distance']:.0f} mm"),
    ]
    story.append(_metric_table(res_rows))
    story.append(Spacer(1, 6 * mm))

    # ── SECTION 2: INSTALLATION DRAWING ──────────────────────────────────────
    story.append(SectionTitle("2. Installation Drawing", usable_w))
    story.append(Spacer(1, 3 * mm))
    story.append(InstallationDrawing(
        width=usable_w,
        height=72 * mm,
        fov_h=params["fov_x"],
        fov_v=params["fov_y"],
        distance=params["distance"],
        roi_w=params.get("roi_w") if params.get("roi_enabled") else None,
        roi_h=params.get("roi_h") if params.get("roi_enabled") else None,
    ))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        "Drawing is schematic and not to scale. Dashed lines indicate the camera FOV cone. "
        "Red overlay (if shown) indicates the configured ROI zone.",
        S["small"],
    ))
    story.append(Spacer(1, 6 * mm))

    # ── SECTION 3: ROI ANALYSIS ───────────────────────────────────────────────
    if params.get("roi_enabled"):
        story.append(SectionTitle("3. Region of Interest (ROI) Analysis", usable_w, C_RED))
        story.append(Spacer(1, 3 * mm))

        roi_metrics = [
            ("ROI Width",        f"{params['roi_w']:.1f}",  "mm"),
            ("ROI Height",       f"{params['roi_h']:.1f}",  "mm"),
            ("ROI Pixels (H)",   f"{params['roi_px_h']:.0f}", "px"),
            ("ROI Pixels (V)",   f"{params['roi_px_v']:.0f}", "px"),
        ]
        story.append(_four_metric_row(roi_metrics))
        story.append(Spacer(1, 3 * mm))
        roi_rows = [
            ("ROI Width",            f"{params['roi_w']:.1f} mm"),
            ("ROI Height",           f"{params['roi_h']:.1f} mm"),
            ("ROI Pixel Dimensions", f"{params['roi_px_h']:.0f} × {params['roi_px_v']:.0f} px"),
            ("Total ROI Pixels",     f"{params['roi_px_h']*params['roi_px_v']/1e6:.2f} MP"),
            ("Coverage of Full FOV", f"{params['roi_coverage']:.1f} %"),
        ]
        story.append(_metric_table(roi_rows))
        story.append(Spacer(1, 6 * mm))

    # ── SECTION 4: FEATURE DETECTION ─────────────────────────────────────────
    if params.get("detection_enabled"):
        sec_num = 4 if params.get("roi_enabled") else 3
        story.append(SectionTitle(f"{sec_num}. Feature Detection Analysis", usable_w, C_GREEN))
        story.append(Spacer(1, 3 * mm))

        detect_metrics = [
            ("Feature Size",     f"{params['feature_size']:.3f}", "mm"),
            ("Pixels on Feature",f"{params['detect_px']:.1f}",   "px"),
            ("Limit @ 1 px",     f"{params['detect_min_1px']:.4f}", "mm"),
            ("Limit @ 3 px",     f"{params['detect_min_3px']:.4f}", "mm"),
        ]
        story.append(_four_metric_row(detect_metrics))
        story.append(Spacer(1, 3 * mm))

        verdict = params.get("detect_verdict", "ok")
        if verdict == "ok":
            verdict_text = f"✔  DETECTABLE — {params['detect_px']:.1f} pixels cover the {params['feature_size']:.3f} mm feature. Detection is reliable."
            verdict_color = C_GREEN
            verdict_bg    = C_GREEN_L
        elif verdict == "marginal":
            verdict_text = f"⚠  MARGINAL — Only {params['detect_px']:.1f} pixels cover the feature. Detection may be unreliable. Consider shorter distance."
            verdict_color = C_ORANGE
            verdict_bg    = colors.HexColor("#fef7e0")
        else:
            verdict_text = f"✘  NOT DETECTABLE — Feature is sub-pixel ({params['detect_px']:.3f} px). Minimum detectable: {params['detect_min_1px']:.4f} mm (1px), {params['detect_min_3px']:.4f} mm (3px)."
            verdict_color = C_RED
            verdict_bg    = C_RED_L

        verdict_style = ParagraphStyle(
            "Verdict", fontName="Helvetica-Bold", fontSize=9,
            textColor=verdict_color, backColor=verdict_bg,
            borderPad=8, leading=14,
        )
        story.append(Paragraph(verdict_text, verdict_style))
        story.append(Spacer(1, 6 * mm))

    # ── SECTION 5: CAMERA SPECIFICATIONS ─────────────────────────────────────
    sec_num = sum([
        3,
        1 if params.get("roi_enabled") else 0,
        1 if params.get("detection_enabled") else 0,
    ])
    story.append(SectionTitle(f"{sec_num}. Camera Specifications", usable_w, C_NAVY))
    story.append(Spacer(1, 3 * mm))

    specs = params.get("camera_specs", {})
    if specs:
        story.append(_metric_table(list(specs.items())))
    story.append(Spacer(1, 6 * mm))

    # ── NOTES ─────────────────────────────────────────────────────────────────
    notes = params.get("notes", "").strip()
    if notes:
        story.append(SectionTitle("Notes", usable_w, C_DKGREY))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(notes, S["body"]))
        story.append(Spacer(1, 6 * mm))

    # ── FOOTER LINE ───────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_LINE))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        f"Keyence IV3 Camera Calculator  ·  Made by Dejan Rožič  ·  {now}",
        ParagraphStyle("Footer", fontName="Helvetica", fontSize=7,
                       textColor=C_DKGREY, alignment=TA_CENTER),
    ))

    doc.build(story)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# EXCEL SESSION LOG
# ════════════════════════════════════════════════════════════════════════════

def generate_excel_log(history: list[dict]) -> bytes:
    """
    Convert session history list of dicts to a formatted Excel workbook.
    Returns bytes.
    """
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, GradientFill, PatternFill, Side
    )

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Session Log"

    if not history:
        ws["A1"] = "No calculations logged yet."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    df      = pd.DataFrame(history)
    headers = list(df.columns)

    # Header style
    hdr_fill   = PatternFill("solid", fgColor="0D1B2A")
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    body_font  = Font(name="Arial", size=9)
    alt_fill   = PatternFill("solid", fgColor="F5F5F5")
    thin_side  = Side(style="thin", color="D0D0D0")
    thin_border= Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Write headers
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = center
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font
            cell.alignment = left
            cell.border    = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Row height
    ws.row_dimensions[1].height = 22
    for i in range(2, len(history) + 2):
        ws.row_dimensions[i].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
